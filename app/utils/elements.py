import tkinter as tk
import os
from tkinter import filedialog
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import column_index_from_string
from . import settings as settings
from openpyxl.styles import Font, Alignment, Border
import requests



class ElementCommand:
    def __init__(self):
        self.files_dict = {'file': None, 'filename': None, 'future_filename': None, 'directory': None}

    def limit_lenght(self, stringvar, textvariable, limit):
        value = stringvar.get()
        if len(value) > limit:
            stringvar.set(value[:limit])
        if textvariable == "entry_input":
            settings.input_column_value = value
        elif textvariable == "entry_output":
            settings.output_column_value = value
        elif textvariable == "entry_name":
            settings.cell_e1_text = value

    def file_selector(self):
        excel_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        excel_filename = os.path.splitext(os.path.basename(excel_file))[0]
        excel_future_name = excel_filename + '_result.xlsx'
        self.inf_saves(excel_file, excel_filename, excel_future_name)

    def directory_selector(self):
        directory = filedialog.askdirectory()
        self.inf_saves(directory=directory)

    def inf_saves(self, file=None, filename=None, future_filename=None, directory=None):
        if file != None:
            self.files_dict["file"] = file
            self.files_dict["filename"] = filename
            self.files_dict["future_filename"] = future_filename
        if directory != None:
            self.files_dict["directory"] = directory

    def process_file(self):
        ws = self.load_excel(self.files_dict["file"])
        input = settings.input_column_value
        output = settings.output_column_value
        if self.verif_columns_entry():
            input, output = self.column_letter_converter(input, output)
            sites = self.verif_cells(ws, input)
            if sites:
                self.write_verification_results(ws, output, sites)
                self.config_output_first_cell(ws, output)
                self.apply_borders(ws, output)
                self.set_column_width(ws, input, output)
                self.save_file(self.files_dict["file"], self.load_file(self.files_dict["directory"], self.files_dict["future_filename"]))
            else:
                pass
        else:
            self.verif_columns_entry()

    def load_excel(self, file):
        wb = load_workbook(file)
        ws = wb.active
        return ws
    
    def verif_columns_entry(self, input, output, sites):
        if input is not None:
            if output is not None:
                if len(input) == 1 and 'A' <= input.upper() <= 'Z':
                    if len(output) == 1 and 'A' <= output.upper() <= 'Z' and output.upper() != input.upper():
                        if sites:
                            return True
                        else: return "MENSAGEM DE ERRO"
                    else: return "MENSAGEM DE ERRO"
                else: return "MENSAGEM DE ERRO"
            else: return "MENSAGEM DE ERRO"
        else: return "MENSAGEM DE ERRO"

    def column_letter_converter(self, input_value, output_value):
        input = input_value.upper()
        output = output_value.upper()
        return input, output
    
    def verif_cells(self, ws, column_value):
        sites = []
        for cell in ws[column_value]:
            if cell.value is not None:
                if cell.value.startswith("http://"):
                    sites.append(cell.value)
                else:
                    sites.append("http://" + cell.value)
        return sites
    
    def write_verification_results(self, ws, output_column, sites):
        column_output = column_index_from_string(output_column)
        with ThreadPoolExecutor(max_workers=10) as executor: 
            results = list(executor.map(self.check_access, sites))
                                    
            for i, result in enumerate(results, start=1):
                if i <= len(sites):
                    ws.cell(row=i, column=column_output).value = result
                    ws.cell(row=i, column=column_output).alignment = settings.alignment_cell
                                            
                    if result == "Accessible":
                        settings.verified_sites += 1
                        ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_acessible)
                    elif result == "Inaccessible":
                        settings.verified_sites += 1
                        ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_inacessible)
                    elif result == "Timeout Exceeded":
                        settings.verified_sites += 1 
                        ws.cell(row=i, column=column_output).font = Font(size=settings.font_size_e, color=settings.font_color_timeout)

    def check_access(self, site):
        try:
            response = requests.get(site, verify=False, timeout=10) 
            return "Accessible" if response.status_code == 200 else "Inaccessible"
        except requests.exceptions.RequestException as e:
            if isinstance(e, requests.exceptions.Timeout):
                return "Timeout Exceeded"
            else:
                return "Inaccessible"

    def config_output_first_cell(self, ws, output_column):
        ws[f'{output_column}1'].value = settings.cell_e1_text
        ws[f'{output_column}1'].font = Font(bold=True, size=settings.first_cell_size)
        ws[f'{output_column}1'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{output_column}1'].border = Border(left=settings.title_border_left, right=settings.title_border_right, top=settings.title_border_top, bottom=settings.title_border_bottom)

    def apply_borders(self, ws, output_column):
        for i in range(2, settings.verified_sites + 1):
            cell = ws.cell(row=i, column=output_column)
            cell.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.cell_border_top, bottom=settings.cell_border_bottom)
    
        cell_2 = ws[f'{output_column}2']
        cell_2.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.end_border_edges_top, bottom=settings.cell_border_bottom)
       
        cell_x = ws[f'{output_column}{settings.verified_sites + 1}']
        cell_x.border = Border(left=settings.cell_border_left, right=settings.cell_border_right, top=settings.cell_border_top, bottom=settings.end_border_edges_bottom)

    def set_column_width(self, ws, input_column, output_column):
        input_column_width = ws.column_dimensions[input_column].width
        ws.column_dimensions[output_column].width = input_column_width

    def save_file(self, file, new_file):
        file.save(new_file)

    def load_file(self, directory, future_filename):
        new_file = os.path.join(directory, future_filename)
        return new_file